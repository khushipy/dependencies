# Standard library imports
import os                    # For file and directory operations
import sys                   # For system-specific functions and command line arguments
import time                  # For time-related functions and delays
import subprocess            # For running external processes
import multiprocessing       # For CPU core detection and parallel processing
from datetime import datetime  # For timestamp generation
from concurrent.futures import ProcessPoolExecutor, wait, FIRST_COMPLETED  # For parallel task execution
import traceback

# Third-party imports
import openpyxl  # For Excel file operations

# Configuration Constants
EXCEL_FILENAME = "input_file.xlsx"  # Default input Excel file name
# Configuration file format:
# Line 1: Number of input columns
# Line 2: Path to the executable to run
# Line 3 (optional): Case ID range in format 'start_id:end_id'
CONFIG_FILENAME = "input.txt"

# Column headers for status tracking in the Excel file
# These columns will be added to track the execution status of each case
STATUS_COL_HEADERS = ["Status", "Start Time", "End Time", "Output"]

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller bundle."""
    import sys
    base_path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base_path, relative_path)


def find_available_cores():
    """
    Determine the number of CPU cores available for processing.
    
    Returns:
        int: Number of available CPU cores, ensuring at least 1 core is available.
             Always leaves 2 cores free for system processes.
    """
    total = multiprocessing.cpu_count()
    return max(1, total - 2)  # Leave 2 cores free for system processes


def load_config(path):
    """
    Load and parse the configuration file.
    
    Args:
        path (str): Path to the configuration file.
        
    Returns:
        tuple: A tuple containing:
            - num_columns (int): Number of input columns in the Excel file
            - exe_name (str): Path to the executable to run
            - case_id_range (tuple): Optional tuple of (start_id, end_id) if specified in config
            
    Raises:
        SystemExit: If the configuration file is invalid or cannot be read
    """
    try:
        # Read all non-empty lines from the config file
        with open(path, "r") as f:
            lines = [line.strip() for line in f if line.strip()]
            
        # Verify minimum required lines exist
        if len(lines) < 2:
            raise ValueError(f"The config file '{path}' must have at least two lines: input column count and exe path")
        
        # Parse required configuration values
        num_columns = int(lines[0])  # First line: number of input columns
        exe_name = lines[1]          # Second line: path to the exe
        
        # Parse optional case ID range (format: "start_id:end_id") eg. 20:30 Inclusive [includes 20 and 30]
        case_id_range = None
        if len(lines) >= 3 and ':' in lines[2]:
            try:
                start, end = lines[2].split(':')
                # Preserve the exact string values for display but convert to int for range generation
                case_id_range = (start.strip(), end.strip())
                print(f"[INFO] Processing case IDs from {start} to {end}")
            except ValueError:
                print(f"[WARNING] Invalid case ID range format in config. Expected 'start_id:end_id', got '{lines[2]}'")
                
        return num_columns, exe_name, case_id_range
        
    except Exception as e:
        print(f"[ERROR] Unable to read config from {path}: {e}")
        sys.exit(1)


def ensure_status_columns(ws, start_col):
    """
    Ensure that status tracking columns exist in the worksheet.
    
    Args:
        ws (Worksheet): The Excel worksheet to modify
        start_col (int): The starting column index where status columns should be added
        
    Returns:
        dict: A dictionary mapping column names to their 1-based indices:
            - status: Status of the case (pending/running/completed)
            - start_time: When the case started processing
            - end_time: When the case completed processing
            - core: Which CPU core was used
            - output: Output from the process
            - next_after_status: Column index after all status columns
    """
    # Create the fixed status columns starting at start_col
    for i, col_name in enumerate(STATUS_COL_HEADERS):
        ws.cell(row=1, column=start_col + i, value=col_name)
    
    # Return a dictionary with column mappings for easy reference
    return {
        "status": start_col,
        "start_time": start_col + 1,
        "end_time": start_col + 2,
        "output": start_col + 3,
        "next_after_status": start_col + len(STATUS_COL_HEADERS)
    }


def reset_running_on_resume(ws, status_col):
    """
    Reset any 'running' statuses to 'pending' when resuming processing.
    
    This is important for recovery if the script was interrupted while processing.
    
    Args:
        ws (Worksheet): The Excel worksheet to update
        status_col (int): The 1-based column index containing the status values
    """
    # Iterate through all rows (starting from row 2 to skip header)
    for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
        # Reset any non-completed status to 'pending'
        if str(row[0].value).strip().lower() != "completed":
            row[0].value = "pending"




def run_exe_on_batch(inputs, exe_name):
    """
    Execute an external program with the given input parameters.
    
    Args:
        inputs (list): List of input parameters to pass to the executable
        exe_name (str): Path to the executable to run
        
    Returns:
        tuple: (success, error_message)
            - success (bool): True if execution was successful, False otherwise
            - error_message (str): Empty string if successful, otherwise contains error details
    """
    try:
        # Convert to absolute path to avoid path resolution issues
        exe_full_path = resource_path(exe_name)
        
        # Prepare command line arguments - convert all inputs to strings
        args = [exe_full_path] + [str(x) for x in inputs]
        
        # Execute the command and wait for it to complete
        # check=True will raise CalledProcessError if return code is non-zero
        creation_flags = 0
        if sys.platform == "win32":
            creation_flags = subprocess.CREATE_NEW_CONSOLE
        result = subprocess.run(args, check=True, creationflags=creation_flags)
        
        # Return success with no error message
        return True, ""
        
    except subprocess.CalledProcessError as e:
        # The process returned a non-zero exit status
        error_msg = f"Process failed with exit code {e.returncode}"
        print(f"[ERROR] {error_msg}: {e}")
        return False, error_msg
        
    except Exception as e:
        # Handle any other exceptions that might occur
        error_msg = f"Unexpected error: {str(e)}"
        print(f"[ERROR] {error_msg}")
        tb = traceback.format_exc()
        print(f"[TRACEBACK] {tb}")
        return False, error_msg
        
        return False, f"Exception: {e}\n{tb}"


def read_output_file(case_id):
    """Read the OutputCaseID.txt file generated by the exe"""
    import sys, os
    # Use folder where main.exe is located
    base_path = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.abspath(os.path.dirname(__file__))
    output_filename = os.path.join(base_path, f"Output{case_id}.txt")
    if not os.path.exists(output_filename):
        return "", False  # No output file found
    try:
        with open(output_filename, "r", encoding="utf-8") as f:
            line = f.readline().rstrip('\n\r')
        return line, True
    except Exception as e:
        print(f"[ERROR] Failed to read output file {output_filename}: {e}")
        return "", False



def process_batch(row_idx, batch_inputs, exe_name, case_id=None):
    """
    Process a single batch of input data by executing an external program and collecting results.
    
    This function is designed to be called in parallel for multiple batches. It handles the execution
    of an external program, captures its output, and returns the processing results in a structured format.
    
    Args:
        row_idx (int): The starting row index of this batch in the Excel worksheet.
        batch_inputs (list): List of input parameters for the external program.
        exe_name (str): Name/path of the external executable to run.
        case_id (str, optional): Unique identifier for this case. If not provided, will be extracted
                              from batch_inputs[11] or generated as "UNKNOWN_{row_idx}".
    
    Returns:
        tuple: A 7-element tuple containing:
            - row_idx (int): The input row index
            - status (str): One of:
                - "completed": Execution successful with output file
                - "p-error": Partial error - execution failed but output file exists
                - "error": Execution failed with no output
            - start_time (str): Timestamp when processing started
            - end_time (str): Timestamp when processing completed
            - output (str): Contents of the output file if available, else empty string
            - error_msg (str): Error message if execution failed, else empty string
    
    The function performs the following steps:
    1. Determines the case ID if not provided
    2. Records the start time
    3. Executes the external program with the batch inputs
    4. Records the end time
    5. Attempts to read the output file
    6. Returns appropriate status and results based on execution success and output file existence
    """
    # If case_id is not provided, try to get it from batch_inputs (position 11)
    # If not available or empty, generate a fallback ID using row index
    if case_id is None:
        case_id = batch_inputs[11] if len(batch_inputs) > 11 and batch_inputs[11] else f"UNKNOWN_{row_idx}"
    
    # Record the start time of processing
    start = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Execute the external program with the batch inputs
    success, error_msg = run_exe_on_batch(batch_inputs, exe_name)
    
    # Record the end time of processing
    end = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    time.sleep(0.5)  # Wait for file to be written

    # Attempt to read the output file for this case
    output_line, file_exists = read_output_file(case_id)
   
    # Determine the appropriate status and return values based on execution results
    if success:
        if file_exists:
            # Successful execution with output file
            return (row_idx, "completed", start, end, output_line, "")
        else:
            # Successful execution but no output file found - this is an error condition
            return (row_idx, "error", start, end, "", 
                   "No output file found after successful execution")
    else:
        # Error case during execution
        if file_exists:
            # Partial error: Execution failed but output file exists
            return (row_idx, "p-error", start, end, output_line, error_msg)
        else:
            # Complete error: Execution failed with no output
            return (row_idx, "error", start, end, "", error_msg)


def is_excel_file_open(filepath):
    """
    Check if an Excel file is currently open by another process.
    
    Args:
        filepath (str): Path to the Excel file to check
        
    Returns:
        bool: True if the file is open/locked by another process, False otherwise
    """
    try:
        # Try to open the file in append mode
        with open(filepath, 'a') as f:
            pass  # File is not locked
        return False
    except IOError:
        # File is locked by another process
        return True

def kill_excel():
    """Forcefully close Excel after 60 seconds."""
    time.sleep(60)
    try:
        if platform.system() == 'Windows':
            subprocess.run(
                ['taskkill', '/f', '/im', 'EXCEL.EXE'],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                shell=True
            )
        print("[INFO] Excel has been force closed.")
    except Exception as e:
        print(f"[WARNING] Could not close Excel: {e}")
        return False
import ctypes, threading, time, subprocess, os, sys

def message_box_async(text, caption):
    """Show a non-blocking Windows MessageBox (safe from any thread)."""
    def _worker():
        MB_OK = 0x0
        MB_ICONWARNING = 0x30
        MB_TOPMOST = 0x40000
        ctypes.windll.user32.MessageBoxW(None, text, caption, MB_OK | MB_ICONWARNING | MB_TOPMOST)
    threading.Thread(target=_worker, daemon=True).start()

def schedule_excel_kill(delay=60):
    """Kill Excel after `delay` seconds, in the background."""
    def _worker():
        time.sleep(delay)
        try:
            subprocess.run(
                ["taskkill", "/f", "/im", "EXCEL.EXE"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=True
            )
            print("[INFO] Excel closed automatically.")
        except Exception as e:
            print(f"[WARNING] Could not close Excel: {e}")
    threading.Thread(target=_worker, daemon=True).start()

def is_excel_file_open(filepath):
    """True if file is locked by Excel (or another process)."""
    try:
        with open(filepath, "a"):
            return False
    except OSError:
        return True
        
def safe_save(wb, excel_path):
    wait_excel(excel_path, timeout=60)  # show popup + auto-close timer
    wb.save(excel_path)


def wait_excel(filepath, timeout=60):
    """
    If the Excel file is open, show a warning popup immediately and
    schedule an auto-close after `timeout` seconds. Then poll until free.
    """
    # If not locked, return quickly
    if not is_excel_file_open(filepath):
        return

    # Alert once and schedule auto-kill
    message_box_async(
        "The Excel file is open. Please close it.\n"
        f"Auto-closing Excel in {timeout} seconds.",
        "Excel is open"
    )
    schedule_excel_kill(timeout)

    # Poll until the file becomes available
    while is_excel_file_open(filepath):
        time.sleep(1)
    print("[INFO] Excel is now free.")

def kill_pythonwexe():
    """
    Forcefully terminate all running pythonw.exe processes.
    
    This function is called during cleanup to ensure no pythonw.exe processes
    remain running after the script completes or encounters an error.
    
    Returns:
        bool: True if the operation was successful, False otherwise.
    """
    import subprocess
    import platform
    import time

    try:
        if platform.system() == 'Windows':
            # For Windows
            subprocess.run(['taskkill', '/f', '/im', 'pythonw.exe'], 
                         stdout=subprocess.DEVNULL, 
                         stderr=subprocess.DEVNULL)
        
        # Give the system a moment 
        time.sleep(2)
        return True
    except Exception as e:
        print(f"[WARNING] Could not close pythonw.exe: {e}")
        return False


def main():
    """
    Main function that orchestrates the batch processing of Excel data.
    
    Execution Flow:
    =============
    1. INITIALIZATION PHASE
       - Set up file paths and validate input files
       - Load configuration and system resources
    
    2. EXCEL SETUP PHASE
       - Load and prepare Excel workbook
       - Initialize status tracking columns
    
    3. BATCH PROCESSING PHASE
       - Process data in parallel using multiple CPU cores
       - Manage process lifecycle and error handling
       - Update Excel with processing results
    """
    # ============================================
    # 1. INITIALIZATION PHASE
    # ============================================
    # 1.1 Get the absolute path of the directory containing this script
    base_dir = os.path.abspath(os.path.dirname(__file__))
    
    # 1.2 Construct full paths to required files
    runtime_dir = (
        os.path.dirname(sys.executable)
        if getattr(sys, 'frozen', False)
        else os.path.abspath(os.path.dirname(__file__))
    ) #For exe to configure path correctly
    excel_path = os.path.join(runtime_dir, EXCEL_FILENAME)
    config_path = os.path.join(runtime_dir, CONFIG_FILENAME)


    # 1.3 Validate that required input files exist
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel file not found: {excel_path}")
        sys.exit(1)

    if not os.path.exists(config_path):
        print(f"[ERROR] Config file not found: {config_path}")
        sys.exit(1)

    # 1.4 Load configuration and determine system resources
    #     - input_col_count: Number of columns containing input data
    #     - exe_name: Path to the executable that processes each batch
    #     - case_id_range: Optional range of case IDs to process (if specified in config)
    input_col_count, exe_name, case_id_range = load_config(config_path)
    
    # 1.5 Determine optimal number of CPU cores to use
    #     (Leaves 2 cores free for system processes)
    n_cores = find_available_cores()

    # 1.6 Log processing parameters for monitoring
    print(f"[INFO] Using {n_cores} CPU cores (2 kept idle).")
    print(f"[INFO] Processing first {input_col_count} columns per batch.")
    print(f"[INFO] Executable: {exe_name}")

    # ============================================
    # 2. EXCEL SETUP PHASE
    # ============================================
    # 2.1 Open and prepare the Excel workbook
    #     - Load the workbook in read-write mode
    #     - Access the active worksheet for processing
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 2.2 Set up status tracking columns
    #     - Adds status columns after the input data columns
    #     - Returns a dictionary with column indices for:
    #       * status: Current status of each row (pending/running/complete/error)
    #       * start_time: When processing began
    #       * end_time: When processing completed
    #       * core: Which CPU core processed the row
    #       * output: Results from processing
    status_cols = ensure_status_columns(ws, input_col_count + 1)
    status_col = status_cols["status"]
    output_start_col = status_cols["next_after_status"]

    # Reset any 'running' statuses to 'pending' in case of a previous interruption
    reset_running_on_resume(ws, status_col)
    wb.save(excel_path)

    # ============================================
    # 3. BATCH PREPARATION PHASE
    # ============================================
    # 3.1 Read all rows from Excel (skipping header row)
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False))
    
    # 3.2 Prepare batches for processing
    #     - Each batch contains row index and its input data
    batches = []
    for i, row in enumerate(rows):
        # 3.2.1 Extract input values from the row
        #       - Convert all values to strings, handle None values
        #       - Only take the first input_col_count columns as specified in config
        batch_inputs = [str(row[j].value) if j < input_col_count and row[j].value is not None else "" 
                       for j in range(input_col_count)]
        excel_row_num = i + 2  # +2 because Excel rows are 1-based and we skipped header
        
        # 3.3 Filter batches based on case ID range (if specified in config)
        #     This allows processing only a specific range of case IDs
        include_batch = True
        if case_id_range:
            # Get case ID from 12th column (index 11)
            case_id_str = batch_inputs[11] if len(batch_inputs) > 11 and batch_inputs[11] else f"UNKNOWN_{excel_row_num}"
            
            # 3.3.1 Try numeric comparison first (if case IDs are numbers)
            try:
                case_id = int(case_id_str) if case_id_str.isdigit() else case_id_str
                start_id = int(case_id_range[0]) if case_id_range[0].isdigit() else case_id_range[0]
                end_id = int(case_id_range[1]) if case_id_range[1].isdigit() else case_id_range[1]
                include_batch = start_id <= case_id <= end_id
            except (ValueError, AttributeError):
                # 3.3.2 Fall back to string comparison if numeric conversion fails
                include_batch = case_id_range[0] <= case_id_str <= case_id_range[1]
            
        if include_batch:
            batches.append((excel_row_num, batch_inputs))

    

    # ============================================
    # 4. BATCH PROCESSING PHASE
    # ============================================
    # 4.1 Filter batches to only include those with 'pending' status
    #     This allows resuming interrupted processing
    pending_batches = [(r, b) for (r, b) in batches 
                      if str(ws.cell(row=r, column=status_col).value).strip().lower() == "pending"]

    # 4.2 Initialize tracking variables
    futures = {}           # Tracks running futures and their row indices
    submitted_indices = set()  # Tracks which batch indices have been submitted
    next_to_submit = 0     # Index of the next batch to submit

    # 4.3 Create a process pool with specified number of worker processes
    with ProcessPoolExecutor(max_workers=n_cores) as executor:
        # 4.4 Submit initial batch of tasks (one per core)
        while next_to_submit < min(n_cores, len(pending_batches)):
            # 4.4.1 Get the next batch to process
            row_idx, batch_inputs = pending_batches[next_to_submit]
            core_no = next_to_submit % n_cores  # Distribute across cores in round-robin

            # 4.4.2 Add a small delay to prevent resource contention
            time.sleep(1)
            
            # 4.4.3 Update Excel with processing status
            start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row_idx, column=status_col).value = "running"
            ws.cell(row=row_idx, column=status_col + 1).value = start_time  # Start time
            safe_save(wb, excel_path)     # optionally add timeout: safe_save(wb, excel_path, timeout=45)

            
            # 4.4.4 Log the case being processed
            case_id = batch_inputs[11] if len(batch_inputs) > 11 and batch_inputs[11] else f"UNKNOWN_{row_idx}"
            print(f"[INFO] CaseID {case_id} executing ")
            
            # 4.4.5 Submit the batch for processing
            fut = executor.submit(process_batch, row_idx, batch_inputs, exe_name, case_id)
            futures[fut] = row_idx  # Track which future corresponds to which row
            submitted_indices.add(next_to_submit)  # Mark this batch as submitted
            next_to_submit += 1  # Move to next batch

        # 4.5 Main processing loop - continues until all batches are complete
        while futures:
            # 4.5.1 Wait for at least one batch to complete
            #       - Uses FIRST_COMPLETED to process results as they become available
            #       - This allows for better load balancing and progress reporting
            done, _ = wait(futures.keys(), return_when=FIRST_COMPLETED)

            # 4.5.2 Check if Excel file is locked by another process
            #       - Prevents file access conflicts
            if is_excel_file_open(excel_path):
                print("[WARNING] The Excel file is open by another process!")
                wait_excel(excel_path)  # Shows popup and waits for user to close the file
                
            # 4.5.3 Process completed batches
            for fut in done:
                row_idx = futures[fut]  # Get the row index for this future
                try:
                    # 4.5.3.1 Get the result of the completed batch
                    #         - r_idx: Row number in Excel
                    #         - status: 'completed' or 'error'
                    #         - start: Processing start time
                    #         - end: Processing end time
                    #         - full_output: Tab-separated output data
                    #         - error_msg: Error message if any
                    r_idx, status, start, end, full_output, error_msg = fut.result()
                    print(f"[DEBUG] Batch result for row {r_idx}: {status}, output={full_output}")

                except Exception as e:
                    # 4.5.3.2 Handle exceptions during batch processing
                    print(f"[ERROR] Exception in batch at row {row_idx}: {e}")
                    ws.cell(row=row_idx, column=status_col).value = "error"
                    safe_save(wb, excel_path)     # optionally add timeout: safe_save(wb, excel_path, timeout=45)
  # Save error status
                    futures.pop(fut)  # Remove completed future
                    continue  # Skip to next completed future

                # 4.5.4 Update Excel with batch results
                # 4.5.4.1 Update status and end time
                ws.cell(row=r_idx, column=status_col).value = status
                ws.cell(row=r_idx, column=status_col + 2).value = end  # End time

                # 4.5.4.2 Process and save output data
                if full_output:
                    # Split tab-separated output into individual fields
                    output_fields = full_output.split('\t')
                    # Write each field to its own column
                    for field_idx, field in enumerate(output_fields):
                        # Calculate target column (status_col + 3 skips status, start, end)
                        col_idx = status_col + 3 + field_idx
                        ws.cell(row=r_idx, column=col_idx).value = field
                else:
                    # Clear output columns if no data was returned
                    ws.cell(row=r_idx, column=status_col + 3).value = ""

                # 4.5.4.3 Save changes to Excel after updating each batch
                safe_save(wb, excel_path)     # optionally add timeout: safe_save(wb, excel_path, timeout=45)

                
                # 4.5.4.4 Remove completed future from tracking
                futures.pop(fut)

                # 4.5.5 Submit next batch if available
                if next_to_submit < len(pending_batches):
                    # 4.5.5.1 Get next batch details
                    next_row_idx, next_batch_inputs = pending_batches[next_to_submit]
                    core_no = next_to_submit % n_cores  # Distribute across cores

                    # 4.5.5.2 Add delay to prevent resource contention
                    time.sleep(1)
                    
                    # 4.5.5.3 Update Excel with new batch status
                    start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=next_row_idx, column=status_col).value = "running"
                    ws.cell(row=next_row_idx, column=status_col + 1).value = start_time
                    safe_save(wb, excel_path)     # optionally add timeout: safe_save(wb, excel_path, timeout=45)

                    
                    # 4.5.5.4 Log the new batch being processed
                    case_id = next_batch_inputs[11] if len(next_batch_inputs) > 11 and next_batch_inputs[11] else f"UNKNOWN_{next_row_idx}"
                    print(f"[INFO] CaseID {case_id} executing")
                    
                    # 4.5.5.5 Submit the new batch for processing
                    f = executor.submit(process_batch, next_row_idx, next_batch_inputs, 
                                      exe_name, case_id)
                    futures[f] = next_row_idx
                    submitted_indices.add(next_to_submit)
                    next_to_submit += 1  # Move to next batch

    # ============================================
    # 5. CLEANUP AND COMPLETION
    # ============================================
    # 5.1 Final save to ensure all changes are written to disk
    wb.save(excel_path)
    print("[INFO] All batches processed successfully.")


# Main entry point
if __name__ == "__main__":
    multiprocessing.freeze_support()
    try:
        main()
    
    except Exception as e:
        print(f"[FATAL ERROR]: {e}")
        traceback.print_exc()
    finally:
        kill_pythonwexe() #kill pythonw.exe


